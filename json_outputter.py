#!/usr/bin/env python2
# json_outputter.py
#
# json_outputter, by Chuan Tan <ct538@cam.ac.uk>
#
# Copyright (C) Chuan Tan 2018
#
# This programme is free software; you may redistribute and/or modify
# it under the terms of the Apache Licence, version 2.0.

# Usage: json_outputter.py [inputfile] [sheetname] [institution_col_id] [country_col_id] [contact_col_id] [ip_col_id]

# Reads an excel file, and processes the IP addresses using standardise-ip.
# It then packages the each row into a JSON object, and ouputs a list of these JSON Objects.
# E.g. JSON Object = {
#       "Institution" : "University of Cambridge"
#       "Country"     : "UK"
#       "Contact"     : "a123@cam.ac.uk"
#       "IP-Range"    : ["123.332.23.2/24","223.225.110.23/32"]
#       }
import sys
from openpyxl import load_workbook
import standardise_ip
import json
import pycountry
import uuid

# Generate dict of country names to country code
countries = {}
for country in pycountry.countries:
    countries[country.name] = country.alpha_2


SHNAME  = "Sheet1"
INS_COL = 1
COU_COL = 2
CON_COL = 3
IPR_COL = 4

def process_file(inputf, sheetname, institution_col_id, country_col_id, contact_col_id, ip_col_id):
    wb = load_workbook(inputf)
    sheet = wb[sheetname]
    row_ids_ip_range = standardise_ip.process(sheet,int(ip_col_id))
    JSON_Objects = []
    for row_id in row_ids_ip_range:
        JSON_Object = {
            "Institution" : sheet.cell(column=int(institution_col_id), row = int(row_id)).value,
            "Country" : sheet.cell(column=int(country_col_id), row = int(row_id)).value,
            "Contact" : sheet.cell(column=int(contact_col_id), row = int(row_id)).value,
            "IP-Range" : row_ids_ip_range[row_id],
            "Country-Code": countries.get(sheet.cell(column=int(country_col_id), row = int(row_id)).value),
            "Institution-uuid" : str(uuid.uuid3(uuid.NAMESPACE_DNS, sheet.cell(column=int(institution_col_id), row = int(row_id)).value.encode("utf-8")))
            }
        JSON_Objects.append(JSON_Object)
    print json.dumps(JSON_Objects)

def run():
    if len(sys.argv) == 2:
        _, inputf = sys.argv
        sheetname = SHNAME
        institution_col_id = INS_COL
        country_col_id = COU_COL
        contact_col_id = CON_COL
        ip_col_id = IPR_COL
    elif len(sys.argv) == 7:
        _, inputf, sheetname, institution_col_id , country_col_id, contact_col_id, ip_col_id = sys.argv
    else:
        print >>sys.stderr, "Not enough arguments!"
        sys.exit()
    process_file(inputf, sheetname, institution_col_id , country_col_id, contact_col_id, ip_col_id)

if __name__ == "__main__":
    run()

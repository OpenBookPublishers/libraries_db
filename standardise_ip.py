#!/usr/bin/env python2

# standardise_ip.py

# standardise-ip, by Chuan Tan <ct538@cam.ac.uk>
#
# Copyright (C) Chuan Tan 2018
#
# This programme is free software; you may redistribute and/or modify
# it under the terms of the Apache Licence, version 2.0.

# Reads from excel spreadsheet and converts them into IPNetwork, then outputs them
# usage: standardise_ip [inputfile] [outputfile]
# You can easily check if an IPAddress is in IPNetwork by using
# If IPAddress in IPNetwork:
#           do whatever


from cidrize import cidrize,CidrizeError
from netaddr import *
from openpyxl import load_workbook
import re
import sys

# Global variables
badones =[]; # Weeding out the bad IP addresses
goodones = []; # Just for record purposes


# Patterns

# 192.168.*.*
pat_privateip = re.compile("192\.168\..*\..*");

# match xx.xxx.xxx.xx-xx.xxx.xx.xx
pat_hyphen = re.compile("\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}-\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}");

# (xx.xx.xx.xx/xx)
pat_bracket = re.compile("\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}/\d{1,2}");

# xx.xxx.xx.*
pat_wildcard = re.compile("\d{1,3}\.\d{1,3}\.\d{1,3}\.\*");

# xx.xxx.xx-xx.*
pat_thirdoctet_wildcard = re.compile("\d{1,3}\.\d{1,3}\.\d{1,3}-\d{1,3}\.\*");

# xx.xxx.xx-xx.xx
pat_thirdoctet = re.compile("\d{1,3}\.\d{1,3}\.\d{1,3}-\d{1,3}\.\d{1,3}");

# EZProxyIP123.405.2.3
pat_words = re.compile("[^0-9.-/*]");

# xx.xxx.(xx-xx).(xx-xx)
pat_two_brackets = re.compile("\d{1,3}\.\d{1,3}\.\(\d{1,3}-\d{1,3}\)\.\(\d{1,3}-\d{1,3}\)")

# xx.xxx.xx.xx
pat_simple = re.compile("\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}")

# xx.xxx.xx.xx-xxx
pat_fourthoctet = re.compile("\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}-\d{1,3}")

# xx.xxx.xxx.x[0-255]
pat_squarebrackets = re.compile("\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1}\[\d{1,3}-\d{1,3}\]");

# xx.xxx.xxx.[xxxx]
pat_squarebrackets_fourth = re.compile("\d{1,3}\.\d{1,3}\.\d{1,3}\.\[\d{1,4}]");




# Dictionary of stuff to replace. For standardising.
def replace_all(text, dic):
    assert type(text) == type(u"")

    for i, j in dic.iteritems():
        text = text.replace(i, j)
    return text

reps = {
    "to": "-",
    "and": "&",
    u"\u002D": "-",
    u"\u058A": "-",
    u"\u05BE": "-",
    u"\u1400": "-",
    u"\u1806": "-",
    u"\u2011": "-",
    u"\u2012": "-",
    u"\u2013": "-",
    u"\u2014": "-",
    u"\u2015": "-",
    u"\u2E3A": "-",
    u"\u2E3B": "-",
    u"\uFE58": "-",
    u"\uFE63": "-",
    u"\uFF0D": "-",
}

def thirdoctet(ip_result):
    prefix = re.search(re.compile("\d{1,3}\.\d{1,3}"), ip_result).group(0)
    suffix = re.split(re.compile("\d{1,3}\.\d{1,3}\.\d{1,3}-\d{1,3}"),
                      ip_result)[1]
    ip_range = re.search(re.compile("\d{1,3}-\d{1,3}"), ip_result).group(0)
    first = ip_range.split("-")[0]
    second = ip_range.split("-")[1]

    if suffix == ".*":
        startip = prefix + "." + first  + ".0"
        endip = prefix + "." + second + ".255"
    else:
        startip = prefix + "." + first  + suffix
        endip = prefix + "." + second  + suffix

    return iprange_to_cidrs(startip, endip)

def two_brackets(ip_result):
    prefix = re.search(re.compile("\d{1,3}\.\d{1,3}"), ip_result).group(0)
    third_octet = re.findall("\d{1,3}-\d{1,3}", ip_result)[0]
    fourth_octet = re.findall("\d{1,3}-\d{1,3}", ip_result)[1]

    third_octet_first = third_octet.split("-")[0]
    third_octet_second = third_octet.split("-")[1]

    fourth_octet_first = fourth_octet.split("-")[0]
    fourth_octet_second = fourth_octet.split("-")[1]

    startip = prefix + "." + third_octet_first + "." + fourth_octet_first
    endip = prefix + "." + third_octet_second + "." + fourth_octet_second

    return iprange_to_cidrs(startip, endip)

def handle_hyphenated_range(ip_result):
    startip, endip = ip_result.split("-")
    return iprange_to_cidrs(startip, endip)

def empty_list(anything):
    return []

class CategorisationError(Exception):
    pass

def categorise(ip): # categorise individual ips
    ip = ip.replace(" ","")

    pattern_handlers = [
        (pat_privateip, empty_list),
        (pat_hyphen, handle_hyphenated_range),
        (pat_bracket, cidrize),
        (pat_thirdoctet_wildcard, thirdoctet),
        (pat_thirdoctet, thirdoctet),
        (pat_wildcard, cidrize),
        (pat_two_brackets, two_brackets),
        (pat_fourthoctet, cidrize),
        (pat_squarebrackets, cidrize),
        (pat_squarebrackets_fourth, cidrize),
        (pat_simple, cidrize)
    ]

    for pattern, handler in pattern_handlers:
        matched = re.search(pattern, ip)
        if matched is not None:
            return handler(matched.group(0))

    # doesn't fit into any defined form. Must be a typo somewhere.
    raise CategorisationError

def screen(rawvalue):
    """Digest the rawvalue of row."""

    def add_dots(s):
        """We are assuming that the input is of the form X,YYY,ZZZ,AAA
           i.e., that the last three bytes are groups of *three* digits
           separated by commas.  If they weren't, they would probably not
           have been misinterpreted as numbers in the first place by Excel."""
        assert s >= 1000000000
        units = [1000000000, 1000000, 1000, 1]
        my_divide = lambda unit : s / unit % 1000
        return u".".join(map(str, map(my_divide, units)))

    if(type(rawvalue)) == type(1L):
        rawvalue = replace_all(add_dots(rawvalue), reps)
    else:
        rawvalue = replace_all(rawvalue, reps)

    # Remove whitespace and produce an array of (hopefully) readable
    # IP Addresses.
    ss = [x.strip() for x in re.split('[,;&]', rawvalue)]

    cidrizedarray = []
    for ip in ss:
        if ip is None or ip is u'':
            continue

        try:
            cidrizedip = categorise(ip)

            for i in cidrizedip:
                if i == None or i is u'':
                    continue
                cidrizedarray.append(i)

        except CidrizeError as e:
            badones.append(ip)
        except CategorisationError as e:
            badones.append(ip)
        except:
            print >>sys.stderr, "Unhandled exception in screen()!"
            raise

    return cidrizedarray

def process(sheet, columnnumber): # process each row
    for i in range(1, sheet.max_row + 1):
        if(sheet.cell(column=columnnumber, row=i).value) is not None:
            rawvalue = sheet.cell(column=columnnumber, row=i).value
            cidrized = screen(rawvalue)
            print >> sys.stderr, "Processed row %d" %(i)
            sheet.cell(column=columnnumber+1, row=i).value = str(cidrized)
            goodones.append(cidrized)

    for ip in badones:
        print ip

    print "There are %d bad ips need fixing." %(len(badones))

# run()
def run():
    _, inputf, outputf, sheetname, columnnumber = sys.argv
    wb = load_workbook(inputf)
    sheet = wb[sheetname]
    process(sheet, columnnumber)
    wb.save(outputf)

if __name__ == "__main__":
    run()
